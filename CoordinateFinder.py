from geopy import Nominatim
from openpyxl import *
from tkinter.filedialog import askopenfilename
from tkinter import Tk

tk =Tk()
tk.withdraw()
file_name = askopenfilename()
wb = load_workbook(filename=file_name)
ws = wb.active
geolocator = Nominatim(user_agent="GreenBlue", timeout= 5)


def getcoor(loc):
    try:
        return loc.latitude, loc.longitude
    except:
        return None,None


for x in range(2,(ws.max_row+1)):
    adr = ws.cell(row=x,column=1).value + " USA"
    location = geolocator.geocode(adr)
    lat, long = getcoor(location)
    lat_cell = ws.cell(row=x,column=2)
    lat_cell.value = lat
    long_cell = ws.cell(row=x,column =3)
    long_cell.value = long


wb.save(file_name)
